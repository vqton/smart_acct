from typing import List, Optional, Dict, Any
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from pydantic import ValidationError as PydanticValidationError

from domain import (
    ChartOfAccounts, AccountType, DCRDirection,
    AccountingRegime, Result, VASValidationError,
)
from infrastructure.repositories.coa_repository import COARepository


class RowValidationError(Exception):
    def __init__(self, row: int, message: str):
        self.row = row
        self.message = message
        super().__init__(f"Row {row}: {message}")


class COAImportUseCase:
    COLUMNS = ["code", "name", "account_type", "drcr_direction",
               "parent_code", "description", "currency", "unit"]

    def __init__(self, session: Session):
        self.repo = COARepository(session)

    def import_from_excel(self, filepath: str, regime: str = "tt99_2025") -> Result:
        path = Path(filepath)
        if not path.exists():
            return Result.failure(FileNotFoundError(f"File not found: {filepath}"))

        if path.suffix.lower() not in (".xlsx", ".xls"):
            return Result.failure(ValueError("Only .xlsx and .xls files are supported"))

        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return Result.failure(ValueError("Excel file is empty"))

        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        col_index = {name: idx for idx, name in enumerate(header) if name}

        missing = [c for c in self.COLUMNS if c not in col_index]
        if missing:
            return Result.failure(
                ValueError(f"Missing required columns: {', '.join(missing)}. "
                           f"Found: {', '.join(header)}")
            )

        data_rows = rows[1:]
        parsed: List[ChartOfAccounts] = []
        errors: List[Dict[str, Any]] = []

        for i, row in enumerate(data_rows, start=2):
            try:
                raw = {col: (str(row[col_index[col]]).strip()
                             if col_index[col] < len(row) and row[col_index[col]] is not None
                             else "")
                       for col in self.COLUMNS}

                if not raw["code"]:
                    continue

                account_type = self._parse_account_type(raw["account_type"], i)
                drcr = self._parse_drcr(raw["drcr_direction"], account_type, i)

                acc = ChartOfAccounts(
                    code=raw["code"],
                    name=raw["name"],
                    account_type=account_type,
                    regime=AccountingRegime(regime),
                    drcr_direction=drcr,
                    parent_code=raw.get("parent_code") or None,
                    description=raw.get("description") or None,
                    currency=(raw.get("currency") or "VND").upper(),
                    unit=(raw.get("unit") or "VND").upper(),
                )
                parsed.append(acc)
            except RowValidationError as e:
                errors.append({"row": e.row, "error": e.message})
            except (VASValidationError, PydanticValidationError) as e:
                errors.append({"row": i, "error": str(e)})

        created: List[str] = []
        import_errors: List[Dict[str, Any]] = list(errors)

        for acc in parsed:
            r = self.repo.create(acc)
            if r.is_success():
                created.append(acc.code)
            else:
                import_errors.append({"row": 0, "code": acc.code, "error": str(r.error)})

        return Result.success({
            "total": len(parsed) + len(errors),
            "created": created,
            "created_count": len(created),
            "errors": import_errors,
            "error_count": len(import_errors),
        })

    def import_from_dicts(self, accounts: List[Dict[str, Any]], regime: str = "tt99_2025") -> Result:
        parsed: List[ChartOfAccounts] = []
        errors: List[Dict[str, Any]] = []

        for i, row in enumerate(accounts, start=1):
            try:
                if not row.get("code"):
                    continue

                account_type = self._parse_account_type(row.get("account_type", ""), i)
                drcr = self._parse_drcr(row.get("drcr_direction", ""), account_type, i)

                acc = ChartOfAccounts(
                    code=row["code"],
                    name=row["name"],
                    account_type=account_type,
                    regime=AccountingRegime(regime),
                    drcr_direction=drcr,
                    parent_code=row.get("parent_code"),
                    description=row.get("description"),
                    currency=(row.get("currency") or "VND").upper(),
                    unit=(row.get("unit") or "VND").upper(),
                )
                parsed.append(acc)
            except KeyError as e:
                errors.append({"row": i, "error": f"Missing field: {e}"})
            except (VASValidationError, RowValidationError, PydanticValidationError) as e:
                errors.append({"row": i, "error": str(e)})

        created: List[str] = []
        import_errors: List[Dict[str, Any]] = list(errors)

        for acc in parsed:
            r = self.repo.create(acc)
            if r.is_success():
                created.append(acc.code)
            else:
                import_errors.append({"code": acc.code, "error": str(r.error)})

        return Result.success({
            "total": len(accounts),
            "created": created,
            "created_count": len(created),
            "errors": import_errors,
            "error_count": len(import_errors),
        })

    def _parse_account_type(self, raw: str, row: int) -> AccountType:
        val = raw.strip().lower().replace(" ", "_")
        try:
            return AccountType(val)
        except ValueError:
            valid = [e.value for e in AccountType]
            raise RowValidationError(row, f"Invalid account_type '{raw}'. Valid: {valid[:10]}...")

    def _parse_drcr(self, raw: str, account_type: AccountType, row: int) -> DCRDirection:
        val = raw.strip().lower()
        if not val:
            return DCRDirection.get_direction(account_type)
        try:
            return DCRDirection(val)
        except ValueError:
            raise RowValidationError(row, f"Invalid drcr_direction '{raw}'. Use 'debit' or 'credit'")
