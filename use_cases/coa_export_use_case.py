from typing import List, Optional, Dict, Any
from io import BytesIO, StringIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from domain import (
    ChartOfAccounts, AccountType, DCRDirection,
    AccountingRegime, AccountStatus, Result,
)
from infrastructure.repositories.coa_repository import COARepository


class COAExportUseCase:
    HEADERS = ["code", "name", "account_type", "drcr_direction",
               "level", "status", "regime", "parent_code",
               "description", "currency", "unit"]

    def __init__(self, session: Session):
        self.repo = COARepository(session)

    def export_to_excel(self, buf: BytesIO, regime: Optional[str] = None,
                        status: Optional[str] = None) -> None:
        accounts = self.repo.list_all(
            regime=AccountingRegime(regime) if regime else None,
            status=AccountStatus(status) if status else None,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Chart of Accounts"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, h in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, acc in enumerate(accounts, start=2):
            ws.cell(row=i, column=1, value=acc.code)
            ws.cell(row=i, column=2, value=acc.name)
            ws.cell(row=i, column=3, value=acc.account_type.value)
            ws.cell(row=i, column=4, value=acc.drcr_direction.value)
            ws.cell(row=i, column=5, value=acc.level)
            ws.cell(row=i, column=6, value=acc.status.value)
            ws.cell(row=i, column=7, value=acc.regime.value)
            ws.cell(row=i, column=8, value=acc.parent_code or "")
            ws.cell(row=i, column=9, value=acc.description or "")
            ws.cell(row=i, column=10, value=acc.currency)
            ws.cell(row=i, column=11, value=acc.unit)

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col) + 2
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max_length, 40)

        wb.save(buf)

    def export_to_csv(self, buf: BytesIO, regime: Optional[str] = None,
                      status: Optional[str] = None) -> None:
        accounts = self.repo.list_all(
            regime=AccountingRegime(regime) if regime else None,
            status=AccountStatus(status) if status else None,
        )

        text_buf = StringIO()
        writer = csv.writer(text_buf)
        writer.writerow(self.HEADERS)
        for acc in accounts:
            writer.writerow([
                acc.code, acc.name, acc.account_type.value,
                acc.drcr_direction.value, acc.level, acc.status.value,
                acc.regime.value, acc.parent_code or "",
                acc.description or "", acc.currency, acc.unit,
            ])
        buf.write(text_buf.getvalue().encode("utf-8"))

    def export_to_json(self, regime: Optional[str] = None,
                       status: Optional[str] = None) -> List[Dict[str, Any]]:
        accounts = self.repo.list_all(
            regime=AccountingRegime(regime) if regime else None,
            status=AccountStatus(status) if status else None,
        )
        return [
            {
                "code": a.code, "name": a.name,
                "account_type": a.account_type.value,
                "drcr_direction": a.drcr_direction.value,
                "level": a.level, "status": a.status.value,
                "regime": a.regime.value, "parent_code": a.parent_code,
                "description": a.description, "currency": a.currency,
                "unit": a.unit,
                "created_at": a.created_at.isoformat() if a.created_at else None,
                "updated_at": a.updated_at.isoformat() if a.updated_at else None,
            }
            for a in accounts
        ]
