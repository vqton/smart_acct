from decimal import Decimal
from flask import jsonify, request

from domain import (
    FinancialStatementType, FSAccountMapping, FSCashFlowMethod,
    ValidationError, VASValidationError,
)
from domain.i18n import ErrorCodes
from presentation import resolve_error
from presentation.fs import fs_bp, _get_session, _json_fs, _json_line_item, _json_mapping
from use_cases.fs import FSUseCases


def _get_uc():
    return FSUseCases(_get_session())


@fs_bp.route("/api/v1/fs/generate", methods=["POST"])
def generate_fs():
    data = request.get_json(silent=True) or {}
    period = data.get("period")
    stype = data.get("statement_type", "B01_DN")
    entity_id = data.get("entity_id", 1)
    generated_by = data.get("generated_by")
    cf_method = data.get("cash_flow_method", "direct")

    if not period:
        return jsonify({"success": False, "error": resolve_error(ErrorCodes.PERIOD_FS_EMPTY)}), 400

    uc = _get_uc()
    try:
        fs_type = FinancialStatementType(stype)
    except ValueError:
        return jsonify({"success": False, "error": resolve_error(ErrorCodes.STATEMENT_TYPE_INVALID)}), 400

    generators = {
        FinancialStatementType.BALANCE_SHEET_GC: uc.generate_b01_dn,
        FinancialStatementType.INCOME_STATEMENT_GC: uc.generate_b02_dn,
        FinancialStatementType.CASH_FLOW_GC: lambda p, e, g: uc.generate_b03_dn(
            p, FSCashFlowMethod(cf_method), e, g),
        FinancialStatementType.NOTES_GC: uc.generate_b09_dn,
    }

    gen = generators.get(fs_type)
    if not gen:
        return jsonify({"success": False, "error": resolve_error(ErrorCodes.FS_UNSUPPORTED_TYPE)}), 400

    result = gen(period, entity_id, generated_by)
    if result.is_failure():
        err = result.get_error()
        return jsonify({"success": False, "error": resolve_error(err.msgid, **err.params)}), 400

    return jsonify({"success": True, "data": _json_fs(result.get_data())}), 201


@fs_bp.route("/api/v1/fs", methods=["GET"])
def list_fs():
    period = request.args.get("period")
    stype = request.args.get("statement_type")
    status = request.args.get("status")
    entity_id = request.args.get("entity_id", type=int)
    limit = request.args.get("limit", 100, type=int)
    offset = request.args.get("offset", 0, type=int)

    uc = _get_uc()
    try:
        fs_type = FinancialStatementType(stype) if stype else None
    except ValueError:
        fs_type = None

    from domain import FSStatus
    fs_status = FSStatus(status) if status else None

    statements = uc.list_statements(period, fs_type, fs_status, entity_id, limit, offset)
    return jsonify({
        "success": True,
        "data": [_json_fs(s) for s in statements],
        "count": len(statements),
    })


@fs_bp.route("/api/v1/fs/<int:fs_id>", methods=["GET"])
def get_fs(fs_id: int):
    uc = _get_uc()
    fs = uc.get_statement(fs_id)
    if not fs:
        return jsonify({"success": False, "error": resolve_error(ErrorCodes.FS_NOT_FOUND)}), 404
    return jsonify({"success": True, "data": _json_fs(fs)})


@fs_bp.route("/api/v1/fs/<int:fs_id>", methods=["DELETE"])
def delete_fs(fs_id: int):
    uc = _get_uc()
    result = uc.delete_statement(fs_id)
    if result.is_failure():
        err = result.get_error()
        return jsonify({"success": False, "error": resolve_error(err.msgid, **err.params)}), 400
    return jsonify({"success": True})


# ── Approval workflow ──────────────────────────────────────────────

@fs_bp.route("/api/v1/fs/<int:fs_id>/submit", methods=["POST"])
def submit_fs(fs_id: int):
    return _approval_action(fs_id, "submit")


@fs_bp.route("/api/v1/fs/<int:fs_id>/review", methods=["POST"])
def review_fs(fs_id: int):
    return _approval_action(fs_id, "review")


@fs_bp.route("/api/v1/fs/<int:fs_id>/approve", methods=["POST"])
def approve_fs(fs_id: int):
    return _approval_action(fs_id, "approve")


@fs_bp.route("/api/v1/fs/<int:fs_id>/sign", methods=["POST"])
def sign_fs(fs_id: int):
    return _approval_action(fs_id, "sign")


@fs_bp.route("/api/v1/fs/<int:fs_id>/reject", methods=["POST"])
def reject_fs(fs_id: int):
    return _approval_action(fs_id, "reject")


@fs_bp.route("/api/v1/fs/<int:fs_id>/amend", methods=["POST"])
def amend_fs(fs_id: int):
    return _approval_action(fs_id, "amend")


def _approval_action(fs_id: int, action: str):
    data = request.get_json(silent=True) or {}
    user = data.get("user", "system")
    reason = data.get("reason")

    uc = _get_uc()
    actions = {
        "submit": uc.submit_fs,
        "review": uc.review_fs,
        "approve": uc.approve_fs,
        "sign": uc.sign_fs,
        "reject": uc.reject_fs,
        "amend": uc.amend_fs,
    }

    handler = actions.get(action)
    if not handler:
        return jsonify({"success": False, "error": "Unknown action"}), 400

    result = handler(fs_id, user, reason)
    if result.is_failure():
        err = result.get_error()
        return jsonify({"success": False, "error": resolve_error(err.msgid, **err.params)}), 400

    return jsonify({"success": True, "data": _json_fs(result.get_data())})


# ── Audit log ──────────────────────────────────────────────────────

@fs_bp.route("/api/v1/fs/<int:fs_id>/audit-log", methods=["GET"])
def get_audit_log(fs_id: int):
    uc = _get_uc()
    logs = uc.get_audit_log(fs_id)
    return jsonify({"success": True, "data": logs})


# ── Export ─────────────────────────────────────────────────────────

@fs_bp.route("/api/v1/fs/<int:fs_id>/export", methods=["GET"])
def export_fs(fs_id: int):
    fmt = request.args.get("format", "html")
    uc = _get_uc()
    result = uc.export_fs(fs_id, fmt)
    if result.is_failure():
        err = result.get_error()
        return jsonify({"success": False, "error": resolve_error(err.msgid, **err.params)}), 400

    data = result.get_data()
    from flask import render_template

    statement_type_names = {
        "B01_DN": "B01_DN", "B02_DN": "B02_DN", "B03_DN": "B03_DN", "B09_DN": "B09_DN",
    }
    tname = statement_type_names.get(data["fs"].statement_type.value, "B01_DN")

    if fmt == "html":
        html = render_template(f"fs/{tname.lower()}.html",
                                fs=data["fs"], entity=data["fs"])
        return html, 200, {"Content-Type": "text/html; charset=utf-8"}
    elif fmt == "pdf":
        html = render_template(f"fs/{tname.lower()}.html",
                                fs=data["fs"], entity=data["fs"])
        try:
            from weasyprint import HTML
            pdf = HTML(string=html).write_pdf()
            return pdf, 200, {"Content-Type": "application/pdf",
                              "Content-Disposition": f"inline; filename={tname}_{data['fs'].period}.pdf"}
        except Exception as e:
            return jsonify({"success": False, "error": resolve_error(ErrorCodes.FS_EXPORT_FAILED)}), 500
    elif fmt == "xlsx":
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = tname
            ws.cell(row=1, column=1, value="Mã số")
            ws.cell(row=1, column=2, value="Chỉ tiêu")
            ws.cell(row=1, column=4, value="Cuối năm")
            ws.cell(row=1, column=5, value="Đầu năm")
            row = 2
            for line in data["fs"].lines:
                ws.cell(row=row, column=1, value=line.ma_so)
                ws.cell(row=row, column=2, value=line.ten_chi_tieu)
                ws.cell(row=row, column=4, value=float(line.current_year))
                if line.previous_year is not None:
                    ws.cell(row=row, column=5, value=float(line.previous_year))
                row += 1
            from io import BytesIO
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.getvalue(), 200, {
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Disposition": f"attachment; filename={tname}_{data['fs'].period}.xlsx",
            }
        except Exception as e:
            return jsonify({"success": False, "error": resolve_error(ErrorCodes.FS_EXPORT_FAILED)}), 500

    return jsonify({"success": False, "error": resolve_error(ErrorCodes.FS_UNSUPPORTED_FORMAT)}), 400


# ── Account mappings ───────────────────────────────────────────────

@fs_bp.route("/api/v1/fs/mappings", methods=["GET"])
def list_mappings():
    stype = request.args.get("statement_type")
    fs_ma_so = request.args.get("fs_ma_so")
    uc = _get_uc()
    fs_type = FinancialStatementType(stype) if stype else None
    mappings = uc.get_mappings(fs_type, fs_ma_so)
    return jsonify({"success": True, "data": [_json_mapping(m) for m in mappings]})


@fs_bp.route("/api/v1/fs/mappings", methods=["POST"])
def create_mapping():
    data = request.get_json(silent=True) or {}
    try:
        mapping = FSAccountMapping(
            fs_ma_so=data.get("fs_ma_so", ""),
            account_code=data.get("account_code", ""),
            weight=Decimal(str(data.get("weight", "1.00"))),
            direction=data.get("direction", "both"),
            statement_type=FinancialStatementType(data.get("statement_type", "B01_DN")),
        )
    except (ValueError, ValidationError) as e:
        return jsonify({"success": False, "error": str(e)}), 400

    uc = _get_uc()
    result = uc.create_mapping(mapping)
    if result.is_failure():
        err = result.get_error()
        return jsonify({"success": False, "error": resolve_error(err.msgid, **err.params)}), 400
    return jsonify({"success": True, "data": _json_mapping(result.get_data())}), 201


@fs_bp.route("/api/v1/fs/mappings/<int:mapping_id>", methods=["DELETE"])
def delete_mapping(mapping_id: int):
    uc = _get_uc()
    result = uc.delete_mapping(mapping_id)
    if result.is_failure():
        err = result.get_error()
        return jsonify({"success": False, "error": resolve_error(err.msgid, **err.params)}), 400
    return jsonify({"success": True})
