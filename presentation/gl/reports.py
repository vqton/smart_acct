from datetime import date
from flask import request, jsonify, render_template, send_file
from decimal import Decimal
from io import BytesIO

from presentation import resolve_error
from presentation.gl import gl_bp, _get_session
from use_cases.gl_use_cases import GLUseCases
from use_cases.gl.templates import (
    generate_journal_template, generate_s01_ledger, generate_subsidiary_template,
    JOURNAL_TYPE_TEMPLATE_MAP, TEMPLATE_NAMES,
)
from domain import ValidationError, JournalType


def _generate_excel(data: dict, report_type: str) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = report_type

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = []
    rows = []

    if report_type in ("trial-balance", "trial_balance"):
        headers = ["Account Code", "Account Name", "Opening Debit", "Opening Credit",
                    "Period Debit", "Period Credit", "Closing Debit", "Closing Credit"]
        for acc in data.get("accounts", []):
            rows.append([
                acc.get("code", ""), acc.get("name", ""),
                str(acc.get("opening_debit", 0)), str(acc.get("opening_credit", 0)),
                str(acc.get("period_debit", 0)), str(acc.get("period_credit", 0)),
                str(acc.get("closing_debit", 0)), str(acc.get("closing_credit", 0)),
            ])
    elif report_type in ("cash-flow", "cash_flow"):
        headers = ["Item", "Code", "Amount"]
        sections = data.get("sections", []) if "sections" in data else data.get("data", {}).get("sections", [])
        for section in sections:
            rows.append([section.get("name", ""), "", ""])
            for item in section.get("items", []):
                rows.append(["  " + item.get("name", ""), item.get("code", ""), str(item.get("amount", 0))])
    elif report_type in ("balance-sheet", "balance_sheet"):
        headers = ["Item", "Code", "Ending Balance", "Opening Balance"]
        sections = data.get("sections", data.get("data", {}).get("sections", []))
        for section in sections:
            rows.append([section.get("name", ""), "", "", ""])
            for item in section.get("items", []):
                rows.append(["  " + item.get("name", ""), item.get("code", ""),
                             str(item.get("ending_balance", 0)), str(item.get("opening_balance", 0))])
    elif report_type in ("income-statement", "income_statement"):
        headers = ["Item", "Code", "This Period", "Cumulative"]
        items = data.get("items", data.get("data", {}).get("items", []))
        for item in items:
            rows.append([item.get("name", ""), item.get("code", ""),
                         str(item.get("this_period", 0)), str(item.get("cumulative", 0))])
    elif report_type in ("journal", "general-ledger", "subsidiary"):
        entries = data.get("entries", [])
        if entries:
            headers = list(entries[0].keys()) if isinstance(entries[0], dict) else ["Field", "Value"]
            for entry in entries:
                if isinstance(entry, dict):
                    rows.append([str(v) for v in entry.values()])

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border

    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for r in range(2, len(rows) + 2):
            cell_val = ws.cell(row=r, column=col).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[chr(64 + col)].width = min(max_len + 3, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@gl_bp.route("/reports/journal/<period>", methods=["GET"])
def generate_journal_report(period):
    journal_type = request.args.get("journal_type", "general")
    format = request.args.get("format", "json")
    company_name = request.args.get("company_name", "")
    address = request.args.get("address", "")

    session = _get_session()
    try:
        uc = GLUseCases(session)
        entries = uc.list_entries(period=period, is_posted=True)
        entries_filtered = [e for e in entries if hasattr(e, 'journal_type') and e.journal_type.value == journal_type]

        jt = JournalType(journal_type)
    except (ValueError, ValidationError) as e:
        return jsonify({"error": resolve_error(e)}), 400
    finally:
        session.close()

    template_code = JOURNAL_TYPE_TEMPLATE_MAP.get(jt, "S03c-DN")
    data = generate_journal_template(template_code, entries_filtered, period, company_name, address)

    if format == "html":
        return render_template("s03_dn_journal.html", **data)
    return jsonify(data)


@gl_bp.route("/reports/general-ledger/<account_code>/<period>", methods=["GET"])
def generate_general_ledger(account_code, period):
    company_name = request.args.get("company_name", "")
    address = request.args.get("address", "")
    format = request.args.get("format", "json")

    session = _get_session()
    try:
        uc = GLUseCases(session)
        entries = uc.list_entries(period=period, is_posted=True, account_id=account_code)
        account_name = account_code
        try:
            from infrastructure.repositories.gl_repository import GLRepository
            from infrastructure.models.coa_models import COAModel
            from sqlalchemy import select
            model = session.execute(
                select(COAModel).where(COAModel.code == account_code)
            ).scalar_one_or_none()
            if model:
                account_name = model.name
        except Exception:
            pass

        opening_balance = uc.get_account_balance(account_code).get("balance", 0)

        data = generate_s01_ledger(
            account_code=account_code, account_name=account_name,
            entries=entries, period=period,
            opening_balance=opening_balance,
            company_name=company_name, address=address,
        )
    except Exception as e:
        return jsonify({"error": resolve_error(e)}), 400
    finally:
        session.close()

    if format == "html":
        return render_template("s01_dn_general_ledger.html", **data)
    return jsonify(data)


@gl_bp.route("/reports/subsidiary/<subsidiary_type>/<period>", methods=["GET"])
def generate_subsidiary_report(subsidiary_type, period):
    company_name = request.args.get("company_name", "")
    address = request.args.get("address", "")
    entity_id = request.args.get("entity_id", type=int)
    format = request.args.get("format", "json")

    template_code = "S06-DN" if subsidiary_type == "ar" else "S05-DN" if subsidiary_type == "ap" else f"S{subsidiary_type.upper()}-DN"

    session = _get_session()
    try:
        uc = GLUseCases(session)
        entries = uc.get_subsidiary_ledger(
            subsidiary_type=subsidiary_type, period=period, entity_id=entity_id,
        )

        data = generate_subsidiary_template(
            template_code=template_code,
            subsidiary_type=subsidiary_type,
            entries=entries,
            period=period,
            company_name=company_name,
            address=address,
        )
    except ValidationError as e:
        return jsonify({"error": resolve_error(e)}), 400
    finally:
        session.close()

    if format == "html":
        return render_template("s05_s06_dn_subsidiary.html", **data)
    return jsonify(data)


@gl_bp.route("/reports/balance-sheet/<period>", methods=["GET"])
def balance_sheet_report(period):
    format = request.args.get("format", "json")
    session = _get_session()
    try:
        uc = GLUseCases(session)
        data = uc.generate_balance_sheet(period)
    finally:
        session.close()
    if format == "html":
        return render_template("trial_balance.html", **data)
    return jsonify(data)


@gl_bp.route("/reports/income-statement/<period>", methods=["GET"])
def income_statement_report(period):
    format = request.args.get("format", "json")
    session = _get_session()
    try:
        uc = GLUseCases(session)
        data = uc.generate_income_statement(period)
    finally:
        session.close()
    return jsonify(data)


@gl_bp.route("/reports/trial-balance/<period>", methods=["GET"])
def trial_balance_report(period):
    format = request.args.get("format", "json")
    session = _get_session()
    try:
        uc = GLUseCases(session)
        data = uc.generate_trial_balance(period)
    finally:
        session.close()
    if format == "html":
        return render_template("trial_balance.html", **data)
    return jsonify(data)


@gl_bp.route("/reports/cash-flow/<period>", methods=["GET"])
def cash_flow_report(period):
    method = request.args.get("method", "direct")
    format = request.args.get("format", "json")
    session = _get_session()
    try:
        uc = GLUseCases(session)
        data = uc.generate_cash_flow(period, method)
    finally:
        session.close()
    if format == "html":
        return render_template("cash_flow.html", **data)
    return jsonify(data)


@gl_bp.route("/reports/export/<report_type>/<period>", methods=["GET"])
def export_report(report_type, period):
    export_format = request.args.get("format", "pdf")
    session = _get_session()
    try:
        uc = GLUseCases(session)
        if report_type == "trial-balance":
            data = uc.generate_trial_balance(period)
            html = render_template("trial_balance.html", **data)
        elif report_type == "cash-flow":
            method = request.args.get("method", "direct")
            data = uc.generate_cash_flow(period, method)
            html = render_template("cash_flow.html", **data)
        elif report_type == "balance-sheet":
            data = uc.generate_balance_sheet(period)
            html = render_template("trial_balance.html", **data)
        else:
            return jsonify({"error": f"Unknown report type: {report_type}"}), 400

        if export_format == "xlsx":
            buf = _generate_excel(data, report_type)
            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"{report_type}_{period}.xlsx",
            )
        if export_format == "pdf":
            from weasyprint import HTML
            pdf_bytes = HTML(string=html).write_pdf()
            return send_file(
                BytesIO(pdf_bytes),
                mimetype="application/pdf",
                as_attachment=True,
                download_name=f"{report_type}_{period}.pdf",
            )
        return html
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        session.close()


@gl_bp.route("/reports/templates", methods=["GET"])
def list_report_templates():
    templates = []
    for code, name in TEMPLATE_NAMES.items():
        templates.append({
            "code": code,
            "name": name,
            "endpoint": f"/api/v1/gl/reports/journal/<period>?journal_type=<type>&format=json",
        })
    return jsonify({"templates": templates, "total": len(templates)})
