from decimal import Decimal
from datetime import datetime, timezone
from pathlib import Path
import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from domain import ChartOfAccounts, AccountType, DCRDirection, AccountingRegime, AccountStatus
from infrastructure.models.coa_models import Base, COAModel, AccountingRegime as DBRegime, AccountStatus as DBStatus
from use_cases.coa_import_use_case import COAImportUseCase
from use_cases.coa import COAUseCases


@pytest.fixture(scope="function")
def session():
    engine = create_engine("sqlite:///:memory:", echo=False)
    Base.metadata.create_all(engine)
    sess = Session(engine)
    yield sess
    sess.close()


def _make_excel(rows: list, tmp_path: Path) -> str:
    wb = Workbook()
    ws = wb.active
    ws.append(["code", "name", "account_type", "drcr_direction",
               "parent_code", "description", "currency", "unit"])
    for r in rows:
        ws.append(r)
    path = tmp_path / "test_import.xlsx"
    wb.save(str(path))
    return str(path)


class TestImportFromDicts:
    def test_import_valid_accounts(self, session):
        uc = COAImportUseCase(session)
        accounts = [
            {"code": "1", "name": "Tài sản", "account_type": "asset", "drcr_direction": "debit"},
            {"code": "1.1", "name": "Tiền mặt", "account_type": "cash", "drcr_direction": "debit"},
        ]
        result = uc.import_from_dicts(accounts)
        assert result.is_success()
        data = result.get_data()
        assert data["created_count"] == 2
        assert "1" in data["created"]
        assert "1.1" in data["created"]
        assert data["error_count"] == 0

    def test_import_duplicate_returns_errors(self, session):
        uc = COAImportUseCase(session)
        repo = COAUseCases(session)
        repo.create_account(code="1", name="Existing", account_type=AccountType.ASSET,
                            drcr_direction=DCRDirection.DEBIT)
        accounts = [
            {"code": "1", "name": "Duplicate", "account_type": "asset", "drcr_direction": "debit"},
            {"code": "2", "name": "New", "account_type": "liability", "drcr_direction": "credit"},
        ]
        result = uc.import_from_dicts(accounts)
        assert result.is_success()
        data = result.get_data()
        assert data["created_count"] == 1
        assert data["error_count"] == 1
        assert data["created"] == ["2"]

    def test_import_invalid_type_returns_error(self, session):
        uc = COAImportUseCase(session)
        accounts = [
            {"code": "1", "name": "Bad", "account_type": "invalid_type", "drcr_direction": "debit"},
        ]
        result = uc.import_from_dicts(accounts)
        assert result.is_success()
        data = result.get_data()
        assert data["created_count"] == 0
        assert data["error_count"] == 1

    def test_import_missing_code_skips_row(self, session):
        uc = COAImportUseCase(session)
        accounts = [
            {"code": "", "name": "No code", "account_type": "asset", "drcr_direction": "debit"},
        ]
        result = uc.import_from_dicts(accounts)
        assert result.is_success()
        data = result.get_data()
        assert data["created_count"] == 0
        assert data["error_count"] == 0

    def test_import_missing_field_returns_error(self, session):
        uc = COAImportUseCase(session)
        accounts = [
            {"code": "1", "name": "No account_type", "drcr_direction": "debit"},
        ]
        result = uc.import_from_dicts(accounts)
        assert result.is_success()
        data = result.get_data()
        assert data["error_count"] == 1

    def test_import_invalid_code_format(self, session):
        uc = COAImportUseCase(session)
        accounts = [
            {"code": "0.1", "name": "Starts with zero", "account_type": "asset", "drcr_direction": "debit"},
        ]
        result = uc.import_from_dicts(accounts)
        assert result.is_success()
        data = result.get_data()
        assert data["created_count"] == 0
        assert data["error_count"] == 1

    def test_import_default_drcr_from_type(self, session):
        uc = COAImportUseCase(session)
        accounts = [
            {"code": "2", "name": "Nợ", "account_type": "liability"},
        ]
        result = uc.import_from_dicts(accounts)
        assert result.is_success()
        data = result.get_data()
        assert data["created_count"] == 1
        fetched = COAUseCases(session).get_account("2").get_data()
        assert fetched.drcr_direction == DCRDirection.CREDIT


class TestImportFromExcel:
    def test_import_excel_valid(self, session, tmp_path):
        uc = COAImportUseCase(session)
        path = _make_excel([
            ["1", "Tài sản", "asset", "debit", "", "", "VND", "VND"],
            ["2", "Nợ phải trả", "liability", "credit", "", "", "VND", "VND"],
        ], tmp_path)
        result = uc.import_from_excel(path)
        assert result.is_success()
        data = result.get_data()
        assert data["created_count"] == 2
        assert data["error_count"] == 0

    def test_import_excel_file_not_found(self, session):
        uc = COAImportUseCase(session)
        result = uc.import_from_excel("/nonexistent/file.xlsx")
        assert result.is_failure()

    def test_import_excel_wrong_extension(self, session):
        uc = COAImportUseCase(session)
        result = uc.import_from_excel("/tmp/test.csv")
        assert result.is_failure()

    def test_import_excel_missing_columns(self, session, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["code", "name"])
        ws.append(["1", "Test"])
        path = tmp_path / "bad_columns.xlsx"
        wb.save(str(path))
        uc = COAImportUseCase(session)
        result = uc.import_from_excel(str(path))
        assert result.is_failure()

    def test_import_excel_empty(self, session, tmp_path):
        wb = Workbook()
        ws = wb.active
        path = tmp_path / "empty.xlsx"
        wb.save(str(path))
        uc = COAImportUseCase(session)
        result = uc.import_from_excel(str(path))
        assert result.is_failure()

    def test_import_excel_partial_errors(self, session, tmp_path):
        uc = COAImportUseCase(session)
        path = _make_excel([
            ["1", "Valid", "asset", "debit"],
            ["0.1", "Invalid starts with 0", "asset", "debit"],
            ["2", "Valid Liability", "liability", "credit"],
        ], tmp_path)
        result = uc.import_from_excel(str(path))
        assert result.is_success()
        data = result.get_data()
        assert data["created_count"] == 2
        assert data["error_count"] == 1

    def test_import_with_parent_code(self, session, tmp_path):
        uc = COAImportUseCase(session)
        path = _make_excel([
            ["1", "Tài sản", "asset", "debit"],
            ["1.1", "Tiền mặt", "cash", "debit", "1"],
        ], tmp_path)
        result = uc.import_from_excel(str(path))
        assert result.is_success()
        data = result.get_data()
        assert data["created_count"] == 2
