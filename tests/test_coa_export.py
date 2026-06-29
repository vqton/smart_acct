from io import BytesIO
import csv
import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from domain import AccountType, DCRDirection, AccountingRegime
from infrastructure.models.coa_models import Base, COAModel, AccountingRegime as DBRegime, AccountStatus as DBStatus
from use_cases.coa_use_cases import COAUseCases
from use_cases.coa_export_use_case import COAExportUseCase


@pytest.fixture(scope="function")
def session():
    engine = create_engine("sqlite:///:memory:", echo=False)
    Base.metadata.create_all(engine)
    sess = Session(engine)
    _seed(sess)
    sess.commit()
    yield sess
    sess.close()


def _seed(sess):
    accounts = [
        COAModel(code="1", name="Tài sản", account_type="asset",
                 regime=DBRegime.TT99_2025, vas_compliant=True,
                 drcr_direction="debit", level=1, status=DBStatus.ACTIVE,
                 currency="VND", unit="VND"),
        COAModel(code="2", name="Nợ phải trả", account_type="liability",
                 regime=DBRegime.TT99_2025, vas_compliant=True,
                 drcr_direction="credit", level=1, status=DBStatus.ACTIVE,
                 currency="VND", unit="VND"),
        COAModel(code="3", name="Vốn chủ sở hữu", account_type="equity",
                 regime=DBRegime.TT99_2025, vas_compliant=True,
                 drcr_direction="credit", level=1, status=DBStatus.ACTIVE,
                 currency="VND", unit="VND"),
    ]
    for a in accounts:
        sess.add(a)


class TestExport:
    def test_export_excel(self, session):
        uc = COAExportUseCase(session)
        buf = BytesIO()
        uc.export_to_excel(buf)
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active
        assert ws.title == "Chart of Accounts"
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 4
        assert rows[0][0] == "code"
        assert rows[1][0] == "1"
        assert rows[2][0] == "2"
        assert rows[3][0] == "3"

    def test_export_excel_filtered_by_status(self, session):
        uc = COAExportUseCase(session)
        buf = BytesIO()
        uc.export_to_excel(buf, status="active")
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 4

    def test_export_csv(self, session):
        uc = COAExportUseCase(session)
        buf = BytesIO()
        uc.export_to_csv(buf)
        buf.seek(0)
        content = buf.getvalue().decode("utf-8")
        reader = csv.DictReader(content.splitlines())
        rows = list(reader)
        assert len(rows) == 3
        assert rows[0]["code"] == "1"
        assert rows[0]["name"] == "Tài sản"

    def test_export_json(self, session):
        uc = COAExportUseCase(session)
        data = uc.export_to_json()
        assert len(data) == 3
        assert data[0]["code"] == "1"
        assert data[0]["account_type"] == "asset"

    def test_export_filtered_by_regime(self, session):
        uc = COAExportUseCase(session)
        data = uc.export_to_json(regime="tt99_2025")
        assert len(data) == 3

    def test_export_empty(self, session):
        repo = COAUseCases(session)
        repo.delete_account("1")
        repo.delete_account("2")
        repo.delete_account("3")
        uc = COAExportUseCase(session)
        data = uc.export_to_json()
        assert data == []
